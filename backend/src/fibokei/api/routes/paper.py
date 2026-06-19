"""Paper trading API endpoints — DB-backed bot state."""

import os
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from sqlalchemy.orm import Session

from fibokei.api.auth import TokenData, get_current_user
from fibokei.api.deps import get_db
from fibokei.db.repository import (
    delete_all_paper_bots,
    delete_paper_bot,
    get_active_paper_bots,
    get_active_phase,
    get_or_create_paper_account,
    get_paper_bot,
    get_paper_bots,
    get_paper_trades,
    get_phase,
    list_phases,
    reset_paper_account,
    reset_paper_counters_and_recalculate,
    save_paper_bot,
    transition_to_new_phase,
    update_paper_bot_state,
)
from fibokei.strategies.registry import strategy_registry

router = APIRouter(tags=["paper"])

# Minimum composite score to promote a combo from research to paper
PROMOTION_THRESHOLD = float(os.environ.get("FIBOKEI_PROMOTION_THRESHOLD", "0.55"))

# Stale-data thresholds: max seconds since last evaluation per timeframe.
# Values are intentionally generous to account for market closures (weekends,
# bank holidays). A bot is only stale if it hasn't seen a bar well beyond
# what a weekend closure would explain.
#   H1  → 26 h (covers overnight + long weekend)
#   H4  → 4 days (covers full weekend + Mon open lag)
#   D   → 14 days (weekly charts, holidays)
STALE_THRESHOLDS = {
    "M1": 300,       # 5 min
    "M5": 1800,      # 30 min
    "M15": 5400,     # 90 min
    "M30": 10800,    # 3 h
    "H1": 93600,     # 26 h
    "H4": 345600,    # 4 days
    "D": 1209600,    # 14 days
}


# ---------- Request / Response schemas ----------

class BotExecutionTargetSpec(BaseModel):
    """Inline target spec for ``POST /paper/bots`` — Phase 2.

    If supplied, the new bot is wired to these execution accounts at
    creation time. Without this list the bot has no targets and (in
    ``db_targets`` mode) defaults to the seeded Paper account.
    """

    execution_account_id: int
    is_enabled: bool = True
    allocation_override: float | None = None
    risk_per_trade_pct_override: float | None = None
    sizing_mode: str = "static_allocation"


class CreateBotRequest(BaseModel):
    strategy_id: str
    instrument: str
    timeframe: str
    risk_pct: float = 1.0
    source_type: str | None = None  # "research" | "backtest" | "manual"
    source_id: str | None = None  # research run_id or backtest id
    # Phase 2: optional explicit execution targets. Empty / omitted →
    # bot defaults to the seeded Paper account in ``db_targets`` mode.
    execution_targets: list[BotExecutionTargetSpec] | None = None


class CreateBotResponse(BaseModel):
    bot_id: str
    strategy_id: str
    instrument: str
    timeframe: str
    state: str
    source_type: str | None = None
    source_id: str | None = None
    execution_targets: list[dict] = []


class BotStatusResponse(BaseModel):
    bot_id: str
    strategy_id: str
    instrument: str
    timeframe: str
    state: str
    bars_seen: int
    has_position: bool
    position: dict | None = None
    last_evaluated_bar: str | None = None
    error_message: str | None = None
    source_type: str | None = None
    source_id: str | None = None


class AccountResponse(BaseModel):
    balance: float
    equity: float
    initial_balance: float
    currency: str = "GBP"
    total_pnl: float
    total_pnl_pct: float
    daily_pnl: float
    weekly_pnl: float
    open_positions: int
    total_trades: int


class BotHealthItem(BaseModel):
    bot_id: str
    strategy_id: str
    instrument: str
    timeframe: str
    state: str
    last_evaluated_bar: str | None
    seconds_since_eval: float | None
    is_stale: bool


class HealthResponse(BaseModel):
    total_bots: int
    active_bots: int
    stale_bots: int
    bots: list[BotHealthItem]


# ---------- Endpoints ----------

@router.post("/paper/bots", response_model=CreateBotResponse)
def create_bot(
    req: CreateBotRequest,
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Create and start a paper trading bot (with promotion gate)."""
    try:
        strategy_registry.get(req.strategy_id)
    except KeyError:
        raise HTTPException(status_code=400, detail=f"Unknown strategy: {req.strategy_id}")

    bot_id = str(uuid.uuid4())[:8]
    source_type = req.source_type or "manual"
    bot_model = save_paper_bot(db, {
        "bot_id": bot_id,
        "strategy_id": req.strategy_id,
        "instrument": req.instrument,
        "timeframe": req.timeframe.upper(),
        "risk_pct": req.risk_pct,
        "source_type": source_type,
        "source_id": req.source_id,
        "state": "monitoring",
    })

    # Phase 2: attach explicit execution targets if supplied.
    target_summaries: list[dict] = []
    if req.execution_targets:
        from fibokei.db.repository import (
            create_bot_execution_target,
            get_execution_account,
        )

        for spec in req.execution_targets:
            acct = get_execution_account(db, spec.execution_account_id)
            if acct is None:
                raise HTTPException(
                    status_code=404,
                    detail=(
                        f"Execution account {spec.execution_account_id} not found"
                    ),
                )
            try:
                target = create_bot_execution_target(
                    db,
                    {
                        "bot_id": bot_id,
                        "execution_account_id": spec.execution_account_id,
                        "is_enabled": spec.is_enabled,
                        "allocation_override": spec.allocation_override,
                        "risk_per_trade_pct_override": spec.risk_per_trade_pct_override,
                        "sizing_mode": spec.sizing_mode,
                    },
                )
            except Exception:
                # Roll back the bot we just created so we don't leave orphans.
                db.rollback()
                raise
            target_summaries.append(
                {
                    "id": target.id,
                    "execution_account_id": target.execution_account_id,
                    "account_name": acct.name,
                    "broker": acct.broker,
                    "environment": acct.environment,
                    "is_enabled": target.is_enabled,
                }
            )

    return CreateBotResponse(
        bot_id=bot_id,
        strategy_id=req.strategy_id,
        instrument=req.instrument,
        timeframe=req.timeframe.upper(),
        state=bot_model.state,
        source_type=source_type,
        source_id=req.source_id,
        execution_targets=target_summaries,
    )


@router.get("/paper/bots", response_model=list[BotStatusResponse])
def list_bots(
    state: str | None = Query(None),
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List all paper trading bots."""
    bots = get_paper_bots(db, state=state)
    return [_bot_to_response(b) for b in bots]


@router.get("/paper/bots/{bot_id}", response_model=BotStatusResponse)
def get_bot_detail(
    bot_id: str,
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get bot detail."""
    bot = get_paper_bot(db, bot_id)
    if not bot:
        raise HTTPException(status_code=404, detail="Bot not found")
    return _bot_to_response(bot)


@router.post("/paper/bots/{bot_id}/stop")
def stop_bot(
    bot_id: str,
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Stop a paper trading bot."""
    bot = update_paper_bot_state(db, bot_id, "stopped")
    if not bot:
        raise HTTPException(status_code=404, detail="Bot not found")
    return {"bot_id": bot_id, "state": bot.state}


@router.post("/paper/bots/{bot_id}/pause")
def pause_bot(
    bot_id: str,
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Pause a paper trading bot."""
    bot = get_paper_bot(db, bot_id)
    if not bot:
        raise HTTPException(status_code=404, detail="Bot not found")
    if bot.state != "monitoring":
        raise HTTPException(status_code=400, detail="Can only pause a monitoring bot")
    updated = update_paper_bot_state(db, bot_id, "paused")
    return {"bot_id": bot_id, "state": updated.state}


@router.post("/paper/bots/{bot_id}/resume")
def resume_bot(
    bot_id: str,
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Resume a paused or stopped paper trading bot (returns to monitoring)."""
    bot = get_paper_bot(db, bot_id)
    if not bot:
        raise HTTPException(status_code=404, detail="Bot not found")
    if bot.state not in ("paused", "stopped"):
        raise HTTPException(status_code=400, detail="Can only resume a paused or stopped bot")
    updated = update_paper_bot_state(db, bot_id, "monitoring")
    return {"bot_id": bot_id, "state": updated.state}


@router.post("/paper/bots/restart-all")
def restart_all_stopped_bots(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Restart all stopped bots — sets their state back to monitoring."""
    stopped = get_paper_bots(db, state="stopped")
    restarted = []
    for bot in stopped:
        update_paper_bot_state(db, bot.bot_id, "monitoring")
        restarted.append(bot.bot_id)
    return {"restarted": restarted, "count": len(restarted)}


class RestoreStaleResponse(BaseModel):
    restored: list[dict]
    needs_attention: list[dict]
    count: int


@router.post("/paper/bots/restore-stale", response_model=RestoreStaleResponse)
def restore_stale_bots(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Recover stale bots in bulk and classify why each is stale.

    - Bots carrying an ``error_message`` (a strategy/adapter exception broke
      them) are *restored*: state→monitoring and the error is cleared, so the
      worker resumes them on its next sync.
    - Bots that are merely *monitoring-but-stale* (no new bar within the
      per-timeframe threshold — typically a data/worker gap) are reported under
      ``needs_attention`` with a reason. They are NOT force-restarted here
      because a running worker holds the bot in memory; the proper fix is the
      worker-side auto-heal (see docs/NEXT_BUILD_STREAMS.md). Cleanly stopped
      bots are left alone (operator intent) — use restart-all for those.
    """
    all_bots = get_paper_bots(db)
    now = datetime.now(timezone.utc)
    restored: list[dict] = []
    needs_attention: list[dict] = []

    for bot in all_bots:
        is_active = bot.state in ("monitoring", "position_open")
        seconds_since = None
        is_stale = False
        if bot.last_evaluated_bar and is_active:
            last_eval = bot.last_evaluated_bar
            if last_eval.tzinfo is None:
                last_eval = last_eval.replace(tzinfo=timezone.utc)
            seconds_since = (now - last_eval).total_seconds()
            threshold = STALE_THRESHOLDS.get(bot.timeframe.upper(), 7200)
            is_stale = seconds_since > threshold

        if bot.error_message:
            prev_state = bot.state
            update_paper_bot_state(db, bot.bot_id, "monitoring", error_message="")
            restored.append({
                "bot_id": bot.bot_id, "instrument": bot.instrument,
                "timeframe": bot.timeframe, "prev_state": prev_state,
                "reason": f"error_cleared: {bot.error_message[:120]}",
            })
        elif is_stale:
            needs_attention.append({
                "bot_id": bot.bot_id, "instrument": bot.instrument,
                "timeframe": bot.timeframe, "reason": "data_or_worker_gap",
                "seconds_since_eval": seconds_since,
                "note": "worker auto-heal required",
            })

    return RestoreStaleResponse(
        restored=restored, needs_attention=needs_attention, count=len(restored)
    )


@router.post("/paper/bots/{bot_id}/restart")
def restart_bot(
    bot_id: str,
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Restart a stopped bot — continues from where it left off (same phase, cumulative PnL)."""
    bot = get_paper_bot(db, bot_id)
    if not bot:
        raise HTTPException(status_code=404, detail="Bot not found")
    if bot.state != "stopped":
        raise HTTPException(status_code=400, detail="Can only restart a stopped bot")
    updated = update_paper_bot_state(db, bot_id, "monitoring")
    return {"bot_id": bot_id, "state": updated.state}


@router.delete("/paper/bots/{bot_id}")
def delete_bot(
    bot_id: str,
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Delete a paper bot and all its trades."""
    deleted = delete_paper_bot(db, bot_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Bot not found")
    return {"deleted": bot_id}


@router.delete("/paper/bots")
def delete_all_bots(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Delete all paper bots and their trades."""
    count = delete_all_paper_bots(db)
    return {"deleted_count": count}


@router.post("/paper/account/reset")
def reset_account(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Reset paper account to initial balance and clear all PnL."""
    acct = reset_paper_account(db)
    return {
        "balance": acct.balance,
        "equity": acct.equity,
        "daily_pnl": acct.daily_pnl,
        "weekly_pnl": acct.weekly_pnl,
        "message": f"Account reset to £{acct.initial_balance:.2f}",
    }


@router.post("/paper/account/reset-counters")
def reset_counters(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Reset daily/weekly PnL counters and recalculate balance from closed trades.

    Use this to correct balance corruption caused by warmup-phase PnL bleed.
    Balance is rebuilt as initial_balance + sum(all closed trade PnL).
    Does NOT wipe trade history or reset balance to zero.
    """
    acct = reset_paper_counters_and_recalculate(db)
    return {
        "balance": acct.balance,
        "equity": acct.equity,
        "daily_pnl": acct.daily_pnl,
        "weekly_pnl": acct.weekly_pnl,
        "message": f"Counters reset. Balance recalculated to £{acct.balance:.2f}",
    }


@router.get("/paper/account", response_model=AccountResponse)
def get_account(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get paper trading account overview."""
    import logging

    logger = logging.getLogger("fibokei.paper")
    try:
        acct = get_or_create_paper_account(db)
    except Exception:
        logger.exception("Failed to get/create paper account")
        raise HTTPException(status_code=500, detail="Failed to load paper account")
    try:
        trades = get_paper_trades(db, limit=10000)
    except Exception:
        logger.exception("Failed to load paper trades")
        trades = []
    try:
        active_bots = get_active_paper_bots(db)
    except Exception:
        logger.exception("Failed to load active bots")
        active_bots = []

    open_count = sum(1 for b in active_bots if b.state == "position_open")
    initial = acct.initial_balance or 1000.0
    currency = getattr(acct, "currency", None) or "GBP"
    total_pnl = (acct.balance or 0.0) - initial
    total_pnl_pct = (total_pnl / initial * 100) if initial > 0 else 0.0
    return AccountResponse(
        balance=acct.balance or 0.0,
        equity=acct.equity or 0.0,
        initial_balance=initial,
        currency=currency,
        total_pnl=total_pnl,
        total_pnl_pct=total_pnl_pct,
        daily_pnl=acct.daily_pnl or 0.0,
        weekly_pnl=acct.weekly_pnl or 0.0,
        open_positions=open_count,
        total_trades=len(trades),
    )


@router.get("/paper/health", response_model=HealthResponse)
def get_health(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Bot health monitoring — stale data detection per bot."""
    all_bots = get_paper_bots(db)
    now = datetime.now(timezone.utc)
    items = []
    stale_count = 0

    for bot in all_bots:
        is_active = bot.state in ("monitoring", "position_open")
        seconds_since = None
        is_stale = False

        if bot.last_evaluated_bar and is_active:
            last_eval = bot.last_evaluated_bar
            if last_eval.tzinfo is None:
                last_eval = last_eval.replace(tzinfo=timezone.utc)
            seconds_since = (now - last_eval).total_seconds()
            threshold = STALE_THRESHOLDS.get(bot.timeframe.upper(), 7200)
            is_stale = seconds_since > threshold

        if is_stale:
            stale_count += 1

        items.append(BotHealthItem(
            bot_id=bot.bot_id,
            strategy_id=bot.strategy_id,
            instrument=bot.instrument,
            timeframe=bot.timeframe,
            state=bot.state,
            last_evaluated_bar=(
                bot.last_evaluated_bar.isoformat() if bot.last_evaluated_bar else None
            ),
            seconds_since_eval=seconds_since,
            is_stale=is_stale,
        ))

    active_count = sum(1 for b in all_bots if b.state in ("monitoring", "position_open"))

    return HealthResponse(
        total_bots=len(all_bots),
        active_bots=active_count,
        stale_bots=stale_count,
        bots=items,
    )


class BotFleetItem(BaseModel):
    bot_id: str
    strategy_id: str
    instrument: str
    timeframe: str
    state: str
    bars_seen: int
    total_trades: int
    total_pnl: float
    has_position: bool
    source_type: str | None = None
    last_evaluated_bar: str | None = None
    is_stale: bool = False


class FleetOverviewResponse(BaseModel):
    total_bots: int
    running: int
    paused: int
    stopped: int
    stale: int
    aggregate_pnl: float
    aggregate_trades: int
    open_positions: int
    bots: list[BotFleetItem]
    strategy_groups: dict[str, dict]


@router.get("/paper/fleet")
def get_fleet_overview(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Fleet-level dashboard: aggregate metrics + per-bot stats.

    All PnL and trade counts are scoped to the current active evaluation phase
    (trades with entry_time >= phase.started_at). If no phase is active, all
    trades are included for backward compatibility.
    """
    all_bots = get_paper_bots(db)
    now = datetime.now(timezone.utc)
    items: list[BotFleetItem] = []
    strategy_groups: dict[str, dict] = {}
    total_pnl = 0.0
    total_trades_count = 0
    open_count = 0

    # Phase-scoped stats: only count trades from the current evaluation phase
    active_phase = get_active_phase(db)
    phase_since = active_phase.started_at if active_phase else None
    # Ensure timezone-aware for comparison
    if phase_since is not None and phase_since.tzinfo is None:
        from datetime import timezone as _tz
        phase_since = phase_since.replace(tzinfo=_tz.utc)

    for bot in all_bots:
        trades = get_paper_trades(db, bot_id=bot.bot_id, limit=10000, since=phase_since)
        bot_pnl = sum(t.pnl for t in trades)
        bot_trades = len(trades)
        total_pnl += bot_pnl
        total_trades_count += bot_trades
        has_position = bot.position_json is not None
        if has_position:
            open_count += 1

        # Stale detection
        is_stale = False
        if bot.last_evaluated_bar and bot.state in ("monitoring", "position_open"):
            last_eval = bot.last_evaluated_bar
            if last_eval.tzinfo is None:
                last_eval = last_eval.replace(tzinfo=timezone.utc)
            seconds_since = (now - last_eval).total_seconds()
            threshold = STALE_THRESHOLDS.get(bot.timeframe.upper(), 7200)
            is_stale = seconds_since > threshold

        items.append(BotFleetItem(
            bot_id=bot.bot_id,
            strategy_id=bot.strategy_id,
            instrument=bot.instrument,
            timeframe=bot.timeframe,
            state=bot.state,
            bars_seen=bot.bars_seen,
            total_trades=bot_trades,
            total_pnl=bot_pnl,
            has_position=has_position,
            source_type=getattr(bot, "source_type", None),
            last_evaluated_bar=(
                bot.last_evaluated_bar.isoformat() if bot.last_evaluated_bar else None
            ),
            is_stale=is_stale,
        ))

        # Strategy grouping
        sid = bot.strategy_id
        if sid not in strategy_groups:
            strategy_groups[sid] = {"count": 0, "running": 0, "pnl": 0.0, "trades": 0}
        strategy_groups[sid]["count"] += 1
        if bot.state in ("monitoring", "position_open"):
            strategy_groups[sid]["running"] += 1
        strategy_groups[sid]["pnl"] += bot_pnl
        strategy_groups[sid]["trades"] += bot_trades

    running = sum(1 for b in all_bots if b.state in ("monitoring", "position_open"))
    paused_count = sum(1 for b in all_bots if b.state == "paused")
    stopped_count = sum(1 for b in all_bots if b.state == "stopped")
    stale_count = sum(1 for i in items if i.is_stale)

    return FleetOverviewResponse(
        total_bots=len(all_bots),
        running=running,
        paused=paused_count,
        stopped=stopped_count,
        stale=stale_count,
        aggregate_pnl=total_pnl,
        aggregate_trades=total_trades_count,
        open_positions=open_count,
        bots=items,
        strategy_groups=strategy_groups,
    )


@router.get("/paper/bots/{bot_id}/trades")
def get_bot_trades(
    bot_id: str,
    limit: int = Query(100, ge=1, le=1000),
    live_only: bool = Query(
        False, description="If true, return only live (forward-monitoring) trades"
    ),
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get trades for a specific bot with equity curve.

    Use live_only=true to filter to trades generated after the bot was created
    (forward-monitoring trades only, excluding historical replay).
    """
    bot = get_paper_bot(db, bot_id)
    if not bot:
        raise HTTPException(status_code=404, detail="Bot not found")
    is_live_filter = True if live_only else None
    trades = get_paper_trades(db, bot_id=bot_id, limit=limit, is_live=is_live_filter)
    # Build equity curve (chronological)
    sorted_trades = sorted(trades, key=lambda t: t.entry_time)
    cumulative = 0.0
    equity_curve = []
    for t in sorted_trades:
        cumulative += t.pnl
        equity_curve.append(round(cumulative, 2))
    return {
        "bot_id": bot_id,
        "total": len(trades),
        "live_only": live_only,
        "trades": [
            {
                "id": t.id,
                "strategy_id": t.strategy_id,
                "instrument": t.instrument,
                "direction": t.direction,
                "entry_time": t.entry_time.isoformat() if t.entry_time else None,
                "entry_price": t.entry_price,
                "exit_time": t.exit_time.isoformat() if t.exit_time else None,
                "exit_price": t.exit_price,
                "exit_reason": t.exit_reason,
                "pnl": t.pnl,
                "bars_in_trade": t.bars_in_trade,
                "is_live": t.is_live,
            }
            for t in trades
        ],
        "equity_curve": equity_curve,
    }


@router.get("/paper/exposure")
def get_exposure(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Portfolio exposure breakdown: per-instrument, per-asset-class, risk utilization."""
    from fibokei.core.instruments import get_instrument
    from fibokei.risk.limits import get_risk_limits

    all_bots = get_paper_bots(db)
    acct = get_or_create_paper_account(db)
    limits = get_risk_limits()
    trades = get_paper_trades(db, limit=10000)

    # Per-instrument exposure (based on open positions)
    instrument_exposure: dict[str, dict] = {}
    active_positions = 0
    for bot in all_bots:
        if bot.position_json is None:
            continue
        active_positions += 1
        inst = bot.instrument
        direction = bot.position_json.get("direction", "LONG")
        if inst not in instrument_exposure:
            instrument_exposure[inst] = {"long": 0, "short": 0, "bot_count": 0}
        instrument_exposure[inst]["bot_count"] += 1
        if direction in ("LONG", "long"):
            instrument_exposure[inst]["long"] += 1
        else:
            instrument_exposure[inst]["short"] += 1

    # Derive net per instrument
    for inst, exp in instrument_exposure.items():
        exp["net"] = exp["long"] - exp["short"]

    # Per-asset-class aggregation
    asset_class_exposure: dict[str, dict] = {}
    for inst, exp in instrument_exposure.items():
        try:
            instrument_obj = get_instrument(inst)
            ac = instrument_obj.asset_class.value
        except KeyError:
            ac = "unknown"
        if ac not in asset_class_exposure:
            asset_class_exposure[ac] = {"long": 0, "short": 0, "instruments": 0}
        asset_class_exposure[ac]["long"] += exp["long"]
        asset_class_exposure[ac]["short"] += exp["short"]
        asset_class_exposure[ac]["instruments"] += 1

    # Concentration warnings: instruments with >= 3 bots
    concentration_warnings = [
        {"instrument": inst, "bot_count": exp["bot_count"]}
        for inst, exp in instrument_exposure.items()
        if exp["bot_count"] >= 3
    ]

    # Risk utilization
    total_long = sum(e["long"] for e in instrument_exposure.values())
    total_short = sum(e["short"] for e in instrument_exposure.values())

    daily_dd_pct = (
        abs(min(acct.daily_pnl, 0.0)) / acct.initial_balance * 100
        if acct.initial_balance > 0 else 0.0
    )
    weekly_dd_pct = (
        abs(min(acct.weekly_pnl, 0.0)) / acct.initial_balance * 100
        if acct.initial_balance > 0 else 0.0
    )

    return {
        "instrument_exposure": instrument_exposure,
        "asset_class_exposure": asset_class_exposure,
        "direction_balance": {"long": total_long, "short": total_short},
        "active_positions": active_positions,
        "concentration_warnings": concentration_warnings,
        "risk_utilization": {
            "open_trades": active_positions,
            # max_open_trades is an emergency backstop — not the primary gate.
            # Primary risk control is drawdown monitoring (daily/weekly limits below).
            "max_open_trades": limits["max_open_trades"],
            "open_trades_pct": (
                round(active_positions / limits["max_open_trades"] * 100, 1)
                if limits["max_open_trades"] > 0 else 0
            ),
            "daily_dd_pct": round(daily_dd_pct, 2),
            "daily_soft_stop_pct": limits["daily_soft_stop_pct"],
            "daily_hard_stop_pct": limits["daily_hard_stop_pct"],
            "weekly_dd_pct": round(weekly_dd_pct, 2),
            "weekly_soft_stop_pct": limits["weekly_soft_stop_pct"],
            "weekly_hard_stop_pct": limits["weekly_hard_stop_pct"],
            # Exposure alert: fires when daily or weekly hard stop is approached
            "drawdown_primary_control": True,
        },
        "total_bots": len(all_bots),
        "total_trades": len(trades),
    }


def _bot_to_response(bot) -> BotStatusResponse:
    """Convert a PaperBotModel to response schema."""
    return BotStatusResponse(
        bot_id=bot.bot_id,
        strategy_id=bot.strategy_id,
        instrument=bot.instrument,
        timeframe=bot.timeframe,
        state=bot.state,
        bars_seen=bot.bars_seen,
        has_position=bot.position_json is not None,
        position=bot.position_json,
        last_evaluated_bar=(
            bot.last_evaluated_bar.isoformat() if bot.last_evaluated_bar else None
        ),
        error_message=bot.error_message,
        source_type=getattr(bot, "source_type", None),
        source_id=getattr(bot, "source_id", None),
    )


# ── Fleet Risk Analysis ─────────────────────────────────────


# ── Evaluation Phases ─────────────────────────────────────────


class PhaseResponse(BaseModel):
    id: int
    name: str
    phase_label: str
    is_active: bool
    started_at: str
    archived_at: str | None = None
    initial_balance: float
    final_balance: float | None = None
    normalized_baseline: float
    broker_balance_at_start: float | None = None
    currency: str
    description: str | None = None
    total_trades: int
    net_pnl: float


class PhaseTransitionRequest(BaseModel):
    # Archive config
    archive_name: str = "Phase A — Initial Testing"
    archive_label: str = "phase_a"
    archive_description: str | None = None
    archive_final_balance: float | None = None
    archive_initial_balance: float = 1000.0
    # New phase config
    new_phase_name: str
    new_phase_label: str
    new_initial_balance: float = 1000.0
    new_normalized_baseline: float = 1000.0
    new_broker_balance: float | None = None
    new_description: str | None = None
    # Bot action: stop all bots before transition?
    stop_active_bots: bool = True
    # Account: reset the paper account balance?
    reset_account: bool = True


def _phase_to_response(p) -> PhaseResponse:
    return PhaseResponse(
        id=p.id,
        name=p.name,
        phase_label=p.phase_label,
        is_active=p.is_active,
        started_at=p.started_at.isoformat() if p.started_at else "",
        archived_at=p.archived_at.isoformat() if p.archived_at else None,
        initial_balance=p.initial_balance,
        final_balance=p.final_balance,
        normalized_baseline=p.normalized_baseline,
        broker_balance_at_start=p.broker_balance_at_start,
        currency=p.currency,
        description=p.description,
        total_trades=p.total_trades or 0,
        net_pnl=p.net_pnl or 0.0,
    )


@router.get("/paper/phases", response_model=list[PhaseResponse])
def list_evaluation_phases(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List all evaluation phases (active + archived)."""
    return [_phase_to_response(p) for p in list_phases(db)]


@router.get("/paper/phases/active", response_model=PhaseResponse | None)
def get_active_evaluation_phase(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get the currently active evaluation phase (or null if none)."""
    p = get_active_phase(db)
    return _phase_to_response(p) if p else None


@router.get("/paper/phases/{phase_id}", response_model=PhaseResponse)
def get_evaluation_phase(
    phase_id: int,
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get a specific evaluation phase by ID."""
    p = get_phase(db, phase_id)
    if not p:
        raise HTTPException(status_code=404, detail="Phase not found")
    return _phase_to_response(p)


@router.post("/paper/phases/transition")
def perform_phase_transition(
    req: PhaseTransitionRequest,
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Archive the current paper trading phase and start a clean new one.

    This operation:
    1. Optionally stops all active bots (so no trades bleed across phases)
    2. Archives all existing bots and trades into the named archive phase
    3. Optionally resets the paper account to the new initial balance
    4. Creates a new active evaluation phase with a clean £1,000 baseline
    """
    import logging
    logger = logging.getLogger("fibokei.paper.phases")

    # 1. Optionally stop all active bots
    if req.stop_active_bots:
        active_bots = get_active_paper_bots(db)
        for bot in active_bots:
            update_paper_bot_state(db, bot.bot_id, "stopped")
        logger.info(
            "Phase transition: stopped %d active bots", len(active_bots)
        )

    # 2. Archive + create new phase
    try:
        archived, new_phase = transition_to_new_phase(
            db,
            new_phase_name=req.new_phase_name,
            new_phase_label=req.new_phase_label,
            archive_name=req.archive_name,
            archive_label=req.archive_label,
            archive_description=req.archive_description,
            archive_final_balance=req.archive_final_balance,
            archive_initial_balance=req.archive_initial_balance,
            new_initial_balance=req.new_initial_balance,
            new_normalized_baseline=req.new_normalized_baseline,
            new_broker_balance=req.new_broker_balance,
            new_description=req.new_description,
        )
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))

    # 3. A new phase always zeroes the period PnL counters (daily/weekly) so
    # the new evaluation starts clean; the worker reloads these on the next
    # cycle via phase-change detection. Rebasing the balance is separate and
    # only happens when the operator asks to reset the account.
    acct = get_or_create_paper_account(db)
    acct.daily_pnl = 0.0
    acct.weekly_pnl = 0.0
    if req.reset_account:
        acct.initial_balance = req.new_initial_balance
        acct.balance = req.new_initial_balance
        acct.equity = req.new_initial_balance
        logger.info(
            "Phase transition: paper account reset to £%.2f", req.new_initial_balance
        )
    db.commit()

    logger.info(
        "Phase transition complete: archived='%s' (id=%d), new='%s' (id=%d)",
        archived.name, archived.id, new_phase.name, new_phase.id,
    )

    return {
        "archived_phase": _phase_to_response(archived),
        "new_phase": _phase_to_response(new_phase),
        "bots_stopped": req.stop_active_bots,
        "account_reset": req.reset_account,
    }


@router.get("/paper/phases/{phase_id}/export")
def export_phase_trades(
    phase_id: int,
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Download all trades from a specific phase as an Excel (.xlsx) file.

    Returns a streaming Excel response. The file contains:
      - Sheet 1: Summary (phase metadata + key metrics)
      - Sheet 2: All trades (chronological)
    """
    import io

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed — cannot export Excel",
        )

    from fastapi.responses import StreamingResponse
    from sqlalchemy import select as _select

    from fibokei.db.models import PaperTradeModel as _PaperTradeModel

    phase = get_phase(db, phase_id)
    if not phase:
        raise HTTPException(status_code=404, detail="Phase not found")

    trades = list(
        db.scalars(
            _select(_PaperTradeModel)
            .where(_PaperTradeModel.phase_id == phase_id)
            .order_by(_PaperTradeModel.entry_time.asc())
        ).all()
    )

    wb = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    def _header(cell_ref, text):
        c = ws_summary[cell_ref]
        c.value = text
        c.font = header_font
        c.fill = header_fill

    def _row(row_num, label, value):
        ws_summary.cell(row=row_num, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_num, column=2, value=value)

    ws_summary.column_dimensions["A"].width = 32
    ws_summary.column_dimensions["B"].width = 30

    _row(1, "Phase Name", phase.name)
    _row(2, "Phase Label", phase.phase_label)
    _row(3, "Status", "Active" if phase.is_active else "Archived")
    _row(4, "Started", phase.started_at.isoformat() if phase.started_at else "")
    _row(5, "Archived", phase.archived_at.isoformat() if phase.archived_at else "—")
    _row(6, "Initial Balance (£)", phase.initial_balance)
    _row(7, "Final Balance (£)", phase.final_balance or "—")
    _row(8, "Normalised Baseline (£)", phase.normalized_baseline)
    _row(9, "IG Broker Balance at Start (£)", phase.broker_balance_at_start or "—")
    _row(10, "Currency", phase.currency)
    _row(11, "Total Trades", phase.total_trades)
    _row(12, "Net PnL (£)", round(phase.net_pnl or 0.0, 2))
    if phase.initial_balance and phase.initial_balance > 0:
        pnl_pct = round((phase.net_pnl or 0.0) / phase.initial_balance * 100, 2)
    else:
        pnl_pct = 0.0
    _row(13, "Net PnL (%)", pnl_pct)

    # Trade-level stats
    if trades:
        winners = [t for t in trades if t.pnl > 0]
        win_rate = round(len(winners) / len(trades) * 100, 1)
        avg_pnl = round(sum(t.pnl for t in trades) / len(trades), 2)
        avg_bars = round(sum(t.bars_in_trade for t in trades) / len(trades), 1)
        best = max(trades, key=lambda t: t.pnl)
        worst = min(trades, key=lambda t: t.pnl)
        _row(15, "Win Rate (%)", win_rate)
        _row(16, "Avg PnL per Trade (£)", avg_pnl)
        _row(17, "Avg Bars in Trade", avg_bars)
        _row(18, "Best Trade PnL (£)", round(best.pnl, 2))
        _row(19, "Worst Trade PnL (£)", round(worst.pnl, 2))
        _row(20, "Description", phase.description or "—")

    # ── Trades sheet ────────────────────────────────────
    ws_trades = wb.create_sheet("Trades")
    trade_headers = [
        "ID", "Bot ID", "Strategy", "Instrument", "Direction",
        "Entry Time", "Entry Price", "Exit Time", "Exit Price",
        "Exit Reason", "PnL (£)", "Bars in Trade",
    ]
    for col_idx, h in enumerate(trade_headers, start=1):
        cell = ws_trades.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    # Cumulative equity curve column
    ws_trades.cell(
        row=1, column=len(trade_headers) + 1, value="Cumulative PnL (£)"
    ).font = Font(bold=True)

    cumulative = 0.0
    for row_idx, t in enumerate(trades, start=2):
        cumulative += t.pnl
        ws_trades.cell(row=row_idx, column=1, value=t.id)
        ws_trades.cell(row=row_idx, column=2, value=t.bot_id)
        ws_trades.cell(row=row_idx, column=3, value=t.strategy_id)
        ws_trades.cell(row=row_idx, column=4, value=t.instrument)
        ws_trades.cell(row=row_idx, column=5, value=t.direction)
        ws_trades.cell(
            row=row_idx, column=6,
            value=t.entry_time.isoformat() if t.entry_time else "",
        )
        ws_trades.cell(row=row_idx, column=7, value=t.entry_price)
        ws_trades.cell(row=row_idx, column=8, value=t.exit_time.isoformat() if t.exit_time else "")
        ws_trades.cell(row=row_idx, column=9, value=t.exit_price)
        ws_trades.cell(row=row_idx, column=10, value=t.exit_reason)
        ws_trades.cell(row=row_idx, column=11, value=round(t.pnl, 2))
        ws_trades.cell(row=row_idx, column=12, value=t.bars_in_trade)
        ws_trades.cell(row=row_idx, column=13, value=round(cumulative, 2))
        # Colour PnL cells
        pnl_cell = ws_trades.cell(row=row_idx, column=11)
        if t.pnl >= 0:
            pnl_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        else:
            pnl_cell.fill = PatternFill("solid", fgColor="FFC7CE")

    # Auto-fit column widths (approximate)
    for col_idx in range(1, len(trade_headers) + 2):
        max_len = max(
            (len(str(ws_trades.cell(r, col_idx).value or "")) for r in range(1, len(trades) + 2)),
            default=10,
        )
        ws_trades.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Stream the file
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"fiboki_phase_{phase.phase_label}_{phase.id}_trades.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/paper/trades/export")
def export_all_trades(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export ALL paper trades (all phases) as an Excel file.

    Useful for the pre-transition archive snapshot. Includes all trades
    regardless of phase assignment, sorted chronologically.
    """
    import io

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed — cannot export Excel",
        )

    from fastapi.responses import StreamingResponse

    trades = get_paper_trades(db, limit=100_000)
    trades_sorted = sorted(trades, key=lambda t: t.entry_time if t.entry_time else datetime.min)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Trades"

    headers = [
        "ID", "Bot ID", "Strategy", "Instrument", "Direction",
        "Entry Time", "Entry Price", "Exit Time", "Exit Price",
        "Exit Reason", "PnL (£)", "Bars in Trade", "Phase ID", "Cumulative PnL (£)",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    cumulative = 0.0
    for row_idx, t in enumerate(trades_sorted, start=2):
        cumulative += t.pnl
        ws.cell(row=row_idx, column=1, value=t.id)
        ws.cell(row=row_idx, column=2, value=t.bot_id)
        ws.cell(row=row_idx, column=3, value=t.strategy_id)
        ws.cell(row=row_idx, column=4, value=t.instrument)
        ws.cell(row=row_idx, column=5, value=t.direction)
        ws.cell(row=row_idx, column=6, value=t.entry_time.isoformat() if t.entry_time else "")
        ws.cell(row=row_idx, column=7, value=t.entry_price)
        ws.cell(row=row_idx, column=8, value=t.exit_time.isoformat() if t.exit_time else "")
        ws.cell(row=row_idx, column=9, value=t.exit_price)
        ws.cell(row=row_idx, column=10, value=t.exit_reason)
        ws.cell(row=row_idx, column=11, value=round(t.pnl, 2))
        ws.cell(row=row_idx, column=12, value=t.bars_in_trade)
        ws.cell(row=row_idx, column=13, value=getattr(t, "phase_id", None))
        ws.cell(row=row_idx, column=14, value=round(cumulative, 2))
        pnl_cell = ws.cell(row=row_idx, column=11)
        if t.pnl >= 0:
            pnl_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        else:
            pnl_cell.fill = PatternFill("solid", fgColor="FFC7CE")

    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(r, col_idx).value or "")) for r in range(1, len(trades_sorted) + 2)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = "fiboki_all_trades_export.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/paper/fleet/risk")
def get_fleet_risk_analysis(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Fleet-level risk analysis: limits status, correlation alerts, underperformers."""
    from fibokei.risk.engine import RiskEngine
    from fibokei.risk.limits import get_risk_limits

    limits = get_risk_limits()
    engine = RiskEngine(**limits)
    all_bots = get_paper_bots(db)

    # Build fleet positions list
    fleet_positions: list[dict] = []
    for bot in all_bots:
        if bot.position_json is not None:
            fleet_positions.append({
                "instrument": bot.instrument,
                "direction": bot.position_json.get("direction", "LONG"),
                "bot_id": bot.bot_id,
            })

    # Per-instrument bot counts (all bots, not just those with positions)
    instrument_bot_counts: dict[str, int] = {}
    for bot in all_bots:
        if bot.state in ("monitoring", "position_open"):
            instrument_bot_counts[bot.instrument] = (
                instrument_bot_counts.get(bot.instrument, 0) + 1
            )

    # Instrument limit breaches
    instrument_alerts = [
        {"instrument": inst, "bot_count": count, "limit": engine.fleet_max_bots_per_instrument}
        for inst, count in instrument_bot_counts.items()
        if count >= engine.fleet_max_bots_per_instrument
    ]

    # Correlation analysis — gather recent trades per bot
    bot_trades: dict[str, list[tuple[str, str]]] = {}
    bot_pnls: dict[str, list[float]] = {}
    for bot in all_bots:
        trades = get_paper_trades(db, bot_id=bot.bot_id, limit=200)
        if trades:
            bot_trades[bot.bot_id] = [
                (
                    t.entry_time.isoformat() if t.entry_time else "",
                    t.exit_time.isoformat() if t.exit_time else "",
                )
                for t in trades
            ]
            bot_pnls[bot.bot_id] = [t.pnl for t in trades]

    correlation_alerts = engine.find_correlated_bots(bot_trades)
    underperformers = engine.find_underperformers(bot_pnls)

    return {
        "fleet_limits": {
            "max_bots_per_instrument": engine.fleet_max_bots_per_instrument,
            "max_total_positions": engine.fleet_max_total_positions,
            "max_exposure_per_instrument": engine.fleet_max_exposure_per_instrument,
            "correlation_threshold": engine.fleet_correlation_threshold,
            "cull_sigma": engine.fleet_cull_sigma,
            "cull_min_trades": engine.fleet_cull_min_trades,
        },
        "fleet_status": {
            "total_bots": len(all_bots),
            "active_bots": sum(1 for b in all_bots if b.state in ("monitoring", "position_open")),
            "open_positions": len(fleet_positions),
            "positions_limit_pct": round(
                len(fleet_positions) / engine.fleet_max_total_positions * 100, 1
            ) if engine.fleet_max_total_positions > 0 else 0,
        },
        "instrument_alerts": instrument_alerts,
        "correlation_alerts": correlation_alerts,
        "underperformers": underperformers,
    }


@router.get("/paper/analytics")
def get_paper_analytics(
    user: TokenData = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Aggregate analytics across all closed paper trades — powers the Analytics dashboard."""
    from collections import defaultdict

    from sqlalchemy import select as sa_select

    from fibokei.db.models import PaperTradeModel

    trades = list(
        db.scalars(
            sa_select(PaperTradeModel).order_by(PaperTradeModel.entry_time.asc())
        ).all()
    )

    empty = {
        "total_trades": 0, "win_trades": 0, "loss_trades": 0,
        "win_rate": 0.0, "avg_win": 0.0, "avg_loss": 0.0,
        "profit_factor": 0.0, "expectancy": 0.0, "total_pnl": 0.0,
        "max_trade_pnl": 0.0, "min_trade_pnl": 0.0,
        "equity_curve": [], "equity_dates": [],
        "pnl_by_strategy": {}, "pnl_by_instrument": {},
        "pnl_by_direction": {}, "pnl_by_exit_reason": {},
        "trade_pnl_list": [], "first_trade_date": None,
        "last_trade_date": None, "days_active": 0,
    }
    if not trades:
        return empty

    wins = [t.pnl for t in trades if t.pnl > 0]
    losses = [t.pnl for t in trades if t.pnl <= 0]
    total = len(trades)
    gross_profit = sum(wins)
    gross_loss = abs(sum(losses))

    cumulative = 0.0
    equity_curve: list[float] = []
    equity_dates: list[str | None] = []
    for t in trades:
        cumulative += t.pnl
        equity_curve.append(round(cumulative, 2))
        equity_dates.append(t.entry_time.isoformat() if t.entry_time else None)

    def make_group() -> dict:
        return {"trades": 0, "pnl": 0.0, "wins": 0}

    by_strategy: dict = defaultdict(make_group)
    by_instrument: dict = defaultdict(make_group)
    by_direction: dict = defaultdict(make_group)
    by_exit_reason: dict = defaultdict(make_group)

    for t in trades:
        for grp, key in [
            (by_strategy, t.strategy_id or "unknown"),
            (by_instrument, t.instrument or "unknown"),
            (by_direction, t.direction or "unknown"),
            (by_exit_reason, t.exit_reason or "unknown"),
        ]:
            grp[key]["trades"] += 1
            grp[key]["pnl"] = round(grp[key]["pnl"] + t.pnl, 2)
            if t.pnl > 0:
                grp[key]["wins"] += 1

    first_date = trades[0].entry_time.isoformat() if trades[0].entry_time else None
    last_date = trades[-1].entry_time.isoformat() if trades[-1].entry_time else None
    days_active = 0
    if trades[0].entry_time and trades[-1].entry_time:
        days_active = (trades[-1].entry_time - trades[0].entry_time).days + 1

    return {
        "total_trades": total,
        "win_trades": len(wins),
        "loss_trades": len(losses),
        "win_rate": round(len(wins) / total * 100, 1) if total > 0 else 0.0,
        "avg_win": round(gross_profit / len(wins), 2) if wins else 0.0,
        "avg_loss": round(sum(losses) / len(losses), 2) if losses else 0.0,
        "profit_factor": round(gross_profit / gross_loss, 2) if gross_loss > 0 else 0.0,
        "expectancy": round(sum(t.pnl for t in trades) / total, 2) if total > 0 else 0.0,
        "total_pnl": round(sum(t.pnl for t in trades), 2),
        "max_trade_pnl": round(max(t.pnl for t in trades), 2),
        "min_trade_pnl": round(min(t.pnl for t in trades), 2),
        "equity_curve": equity_curve,
        "equity_dates": equity_dates,
        "pnl_by_strategy": dict(by_strategy),
        "pnl_by_instrument": dict(by_instrument),
        "pnl_by_direction": dict(by_direction),
        "pnl_by_exit_reason": dict(by_exit_reason),
        "trade_pnl_list": [round(t.pnl, 2) for t in trades],
        "first_trade_date": first_date,
        "last_trade_date": last_date,
        "days_active": days_active,
    }
